import pandas as pd
import statsmodels.formula.api as smf
import numpy as np
import scipy.stats as stats
import warnings
from openpyxl.styles import Font
warnings.filterwarnings("ignore")

import os
script_dir = os.path.dirname(os.path.abspath(__file__))
main_dir = os.path.dirname(os.path.dirname(script_dir))
output_path = os.path.join(main_dir, 'results/aal_contrasts_results.xlsx')
path_aal = os.path.join(main_dir, 'data/AAL/all_data.csv')

try:
    df = pd.read_csv(path_aal)
    if 'ayah' in df['dataset'].values or df['dataset'].str.contains('ayah', case=False).any():
        df = df[~df['dataset'].str.contains('ayah', case=False)]
        
    m_mixed = smf.mixedlm('SampEn ~ dataset', df, groups=df['Subject']).fit()
    
    fe_params = m_mixed.fe_params
    cov_params = m_mixed.cov_params()
    
    fe_names = fe_params.index.tolist()
    beta = fe_params.values
    
    if cov_params.shape[0] > len(beta):
        V = cov_params.loc[fe_names, fe_names].values
    else:
        V = cov_params.values
        
    unique_cats = df['dataset'].unique()
    ref_cat = None
    for cat in unique_cats:
        if f'dataset[T.{cat}]' not in fe_names and cat != 'Intercept':
            ref_cat = cat
            break
            
    def make_contrast_dict(catA, catB):
        d = {}
        if catA != ref_cat:
            d[f'dataset[T.{catA}]'] = 1.0
        if catB != ref_cat:
            d[f'dataset[T.{catB}]'] = -1.0
        return d
        
    # Using correct labels for AAL
    wake_label = 'anestesia_block1'
    lsd_plcb_label = 'lsd_plcb'
    dmt_plcb_label = 'dmt_pcb'
    
    contrasts = [
        ("LSD-placebo vs. Wake", make_contrast_dict(lsd_plcb_label, wake_label)),
        ("DMT-placebo vs. Wake", make_contrast_dict(dmt_plcb_label, wake_label)),
        ("LSD-placebo vs. DMT-placebo", make_contrast_dict(lsd_plcb_label, dmt_plcb_label))
    ]
    
    contrast_results = []
    for c_name, c_dict in contrasts:
        c = np.zeros(len(beta))
        for p_name, weight in c_dict.items():
            if p_name in fe_names:
                c[fe_names.index(p_name)] = weight
                
        estimate = np.dot(c, beta)
        variance = np.dot(c.T, np.dot(V, c))
        stderr = np.sqrt(variance)
        z_stat = estimate / stderr
        p_val = 2 * (1 - stats.norm.cdf(np.abs(z_stat)))
        
        contrast_results.append({
            'Contrast': c_name,
            'Estimate': estimate,
            'Std.Err': stderr,
            'z': z_stat,
            'p-value': p_val
        })
        
    df_contrasts = pd.DataFrame(contrast_results)
    
    # Full model results
    full_results = []
    for name_param in fe_names:
        idx = fe_names.index(name_param)
        est = beta[idx]
        se = np.sqrt(V[idx, idx])
        z = est / se
        p = 2 * (1 - stats.norm.cdf(np.abs(z)))
        full_results.append({
            'Parameter': name_param,
            'Estimate': est,
            'Std.Err': se,
            'z': z,
            'p-value': p
        })
    df_full = pd.DataFrame(full_results)
    
    # Save to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_contrasts.to_excel(writer, sheet_name='AAL', index=False, startrow=1)
        worksheet = writer.sheets['AAL']
        worksheet['A1'] = f"Specific Contrasts (Reference Wake: {wake_label}, Intercept: {ref_cat})"
        
        bold_font = Font(bold=True)
        
        # Format p-values and make bold if significant
        for row in range(3, 3 + len(df_contrasts)):
            p_cell = worksheet.cell(row=row, column=5) # Contrast, Estimate, Std.Err, z, p-value
            p_cell.number_format = '0.000000'
            if p_cell.value is not None and float(p_cell.value) < 0.05:
                p_cell.font = bold_font
                
        # Write full model below
        start_row = len(df_contrasts) + 4
        worksheet.cell(row=start_row, column=1, value="Full Model Fixed Effects")
        df_full.to_excel(writer, sheet_name='AAL', index=False, startrow=start_row)
        
        # Format full model p-values
        for row in range(start_row + 2, start_row + 2 + len(df_full)):
            p_cell = worksheet.cell(row=row, column=5)
            p_cell.number_format = '0.000000'
            if p_cell.value is not None and float(p_cell.value) < 0.05:
                p_cell.font = bold_font
                
    print(f"Excel file saved to: {output_path}")
    
except Exception as e:
    print(f"Error: {e}")
